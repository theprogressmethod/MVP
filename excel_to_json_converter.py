#!/usr/bin/env python3
"""
Excel to JSON Converter for TPM Scoreboards
Converts Excel attendance and commitment data to JSON format for Supabase ingestion
"""

import json
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("Installing required packages...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill


def detect_attendance_from_color(cell):
    """
    Detect attendance status from cell background color
    Green = attended, Red = not_attended, No color = unknown
    """
    if cell.fill and cell.fill.fill_type == 'solid':
        color = cell.fill.start_color.rgb
        if color:
            # Convert to hex if it's not already
            if hasattr(color, 'rgb'):
                hex_color = color.rgb
            else:
                hex_color = str(color).upper()
            
            # Green variations (attendance)
            if any(green in hex_color for green in ['92D050', '00FF00', 'C6EFCE', '90EE90']):
                return "attended"
            # Red variations (no attendance) 
            elif any(red in hex_color for red in ['FF0000', 'FFC7CE', 'FFB3BA']):
                return "not_attended"
    
    return "unknown"


def detect_fulfillment_from_emoji(text):
    """
    Detect commitment fulfillment from emoji
    ✅ = true, ❌ = false, no emoji = unknown
    """
    if not text:
        return "unknown"
    
    text = str(text).strip()
    if text.startswith('✅'):
        return "true"
    elif text.startswith('❌'):
        return "false"
    
    return "unknown"


def clean_commitment_text(text):
    """
    Clean commitment text by removing emoji and extra whitespace
    """
    if not text:
        return ""
    
    text = str(text).strip()
    # Remove leading emoji and whitespace
    text = re.sub(r'^[✅❌]\s*', '', text)
    return text.strip()


def parse_date_cell(cell_value):
    """
    Parse various date formats from Excel cells
    """
    if not cell_value:
        return None
    
    if isinstance(cell_value, datetime):
        return cell_value.strftime('%Y-%m-%d')
    
    # Try to parse string dates
    date_str = str(cell_value).strip()
    if not date_str or date_str.lower() in ['nan', 'none', '']:
        return None
    
    # Common date patterns
    date_patterns = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%d/%m/%Y'
    ]
    
    for pattern in date_patterns:
        try:
            parsed_date = datetime.strptime(date_str, pattern)
            return parsed_date.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    return None


def parse_commitments_column(cell):
    """
    Parse commitments from a cell, splitting by newlines
    """
    if not cell.value:
        return []
    
    commitments = []
    lines = str(cell.value).split('\n')
    
    for line in lines:
        line = line.strip()
        if line:  # Skip empty lines
            fulfillment = detect_fulfillment_from_emoji(line)
            clean_text = clean_commitment_text(line)
            
            if clean_text:  # Only add if there's actual text
                commitments.append({
                    "text": clean_text,
                    "fulfilled": fulfillment
                })
    
    return commitments


def find_data_start_row(worksheet):
    """
    Find the row where actual data starts (after headers)
    Look for the first row with a date
    """
    for row_idx in range(1, min(20, worksheet.max_row + 1)):
        for col_idx in range(1, min(10, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row_idx, col_idx).value
            if parse_date_cell(cell_value):
                return row_idx
    
    return 5  # Default fallback


def find_date_and_commitment_columns(worksheet):
    """
    Find which columns contain dates and commitments
    """
    date_col = None
    commitment_col = None
    
    # Look in first 10 rows and columns for patterns
    for col_idx in range(1, min(10, worksheet.max_column + 1)):
        for row_idx in range(1, min(10, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row_idx, col_idx).value
            
            # Check if this column has dates
            if parse_date_cell(cell_value) and not date_col:
                date_col = col_idx
            
            # Check if this column has commitment-like text
            if (cell_value and isinstance(cell_value, str) and 
                len(str(cell_value)) > 50 and 
                ('✅' in str(cell_value) or '❌' in str(cell_value)) and 
                not commitment_col):
                commitment_col = col_idx
    
    return date_col, commitment_col


def convert_excel_to_json(file_path):
    """
    Main function to convert Excel file to JSON
    """
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    result = {}
    
    print(f"Processing {len(workbook.sheetnames)} sheets...")
    
    for sheet_name in workbook.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        
        # Find key columns
        date_col, commitment_col = find_date_and_commitment_columns(worksheet)
        if not date_col:
            print(f"  Warning: No date column found in {sheet_name}")
            continue
        
        # Find where data starts
        data_start_row = find_data_start_row(worksheet)
        
        user_data = {}
        
        # Process each row of data
        for row_idx in range(data_start_row, worksheet.max_row + 1):
            date_cell = worksheet.cell(row_idx, date_col)
            date_str = parse_date_cell(date_cell.value)
            
            if not date_str:
                continue
            
            # Get attendance from date cell color
            attendance = detect_attendance_from_color(date_cell)
            
            # Get commitments if column exists
            commitments = []
            if commitment_col:
                commitment_cell = worksheet.cell(row_idx, commitment_col)
                commitments = parse_commitments_column(commitment_cell)
            
            user_data[date_str] = {
                "attendance": attendance,
                "commitments": commitments
            }
        
        if user_data:
            result[sheet_name] = user_data
            print(f"  ✅ Processed {len(user_data)} entries for {sheet_name}")
        else:
            print(f"  ⚠️  No data found for {sheet_name}")
    
    return result


def main():
    """
    Main execution function
    """
    input_file = "/Users/thomasmulhern/Downloads/TPM Scoreboards (1).xlsx"
    output_file = "/Users/thomasmulhern/the_progress_method/the_progress_method/tpm_scoreboards_data.json"
    
    if not Path(input_file).exists():
        print(f"Error: File not found - {input_file}")
        return
    
    try:
        print("Converting Excel to JSON...")
        data = convert_excel_to_json(input_file)
        
        # Save to JSON file
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ Conversion complete!")
        print(f"📁 Output saved to: {output_file}")
        print(f"📊 Processed {len(data)} users")
        
        # Print summary
        total_entries = sum(len(user_data) for user_data in data.values())
        total_commitments = sum(
            len(entry["commitments"]) 
            for user_data in data.values() 
            for entry in user_data.values()
        )
        
        print(f"📈 Total entries: {total_entries}")
        print(f"📝 Total commitments: {total_commitments}")
        
    except Exception as e:
        print(f"❌ Error during conversion: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()